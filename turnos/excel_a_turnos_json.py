# -*- coding: utf-8 -*-
"""
excel_a_turnos_json.py (v2 – colores robustos)
Lee bloques "SEMANA" por pestaña (Hotel); E..K=Lunes..Domingo; orden Excel (EmpRow).
Vacaciones: fondo ROJO en C (Nombre). Sustituto: fondo AMARILLO en L.
Cuenta Noches. Extrae color solo si el relleno es sólido y en RGB.

Uso:
  python excel_a_turnos_json.py "TURNOS  v.5.2.xlsx" "turnos_combined_from_excel.json"
Requisitos:
  pip install openpyxl
"""
import sys, json, datetime, argparse
from collections import Counter
from openpyxl import load_workbook

DAY_NAMES = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

def norm_hex_rgb(rgb):
    """Convierte 'AARRGGBB' o 'RRGGBB' a '#rrggbb'. Devuelve '' si no es válido."""
    if not rgb:
        return ""
    s = str(rgb).strip()
    # Si viene como AARRGGBB, nos quedamos con RRGGBB
    if len(s) == 8:  # ej: FFFF0000
        s = s[-6:]
    if len(s) == 6:
        try:
            int(s, 16)
            return "#" + s.lower()
        except Exception:
            return ""
    return ""

def get_cell_hex(cell):
    """
    Devuelve '#rrggbb' del color de fondo SOLO si:
    - el relleno es sólido (patternType='solid'), y
    - el color es de tipo RGB (fgColor.type=='rgb' o start_color.type=='rgb')
    Si es theme/indexed/auto → devuelve '' (mejor usar un color sólido estándar).
    """
    if cell is None or cell.fill is None:
        return ""
    fill = cell.fill
    try:
        if getattr(fill, "patternType", None) != "solid":
            return ""
    except Exception:
        return ""
    # Prioridad: fgColor.rgb con type 'rgb'
    fg = getattr(fill, "fgColor", None)
    if fg is not None and getattr(fg, "type", None) == "rgb":
        return norm_hex_rgb(getattr(fg, "rgb", None))
    # Respaldo: start_color.rgb con type 'rgb'
    sc = getattr(fill, "start_color", None)
    if sc is not None and getattr(sc, "type", None) == "rgb":
        return norm_hex_rgb(getattr(sc, "rgb", None))
    return ""

def is_night(code, longtxt):
    if code == "N":
        return True
    t = (longtxt or "").lower().strip()
    return t in ("noches", "noche")

def parse_excel(path_xlsx):
    wb = load_workbook(path_xlsx, data_only=True)
    rows = []
    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        r = 1
        max_r = sh.max_row or 0
        while r <= max_r:
            c3 = sh.cell(row=r, column=3).value  # C
            if isinstance(c3, str) and c3.strip().upper() == "SEMANA":
                # Nº semana en D
                v_sem = sh.cell(row=r, column=4).value
                try:
                    semana_num = int(v_sem) if v_sem not in (None, "") else None
                except Exception:
                    semana_num = None
                # Fechas E..K
                date_row = r + 1
                day_cols = list(range(5, 12))  # 5..11 = E..K
                day_dates = {}
                for idx, dname in zip(day_cols, DAY_NAMES):
                    val = sh.cell(row=date_row, column=idx).value
                    if isinstance(val, datetime.datetime):
                        day_dates[dname] = val.date().isoformat()
                    elif isinstance(val, datetime.date):
                        day_dates[dname] = val.isoformat()
                week_start = day_dates.get("Lunes", next(iter(day_dates.values()), ""))

                # Filas de empleados
                rr = r + 2
                while rr <= max_r:
                    name_cell = sh.cell(row=rr, column=3)  # C (Nombre)
                    name_val = name_cell.value
                    c1 = sh.cell(row=rr, column=1).value
                    # Fin de bloque
                    if (name_val is None or str(name_val).strip() == "" or str(name_val).strip().upper() == "SEMANA") and (c1 is None or str(c1).strip() == ""):
                        break
                    if name_val is None or str(name_val).strip() == "" or str(name_val).strip().upper() == "SEMANA":
                        rr += 1
                        continue

                    emp = str(name_val).strip()
                    name_hex = get_cell_hex(name_cell)            # C → vacaciones
                    sub_hex  = get_cell_hex(sh.cell(row=rr, column=12))  # L → sustituto

                    for idx, dname in zip(day_cols, DAY_NAMES):
                        cell = sh.cell(row=rr, column=idx)
                        raw = cell.value
                        if raw is None:
                            tlong, code = "", ""
                        else:
                            t = str(raw).strip()
                            low = t.lower()
                            if low in ("m","mañana","manana"):
                                tlong, code = "Mañana", "M"
                            elif low in ("t","tarde"):
                                tlong, code = "Tarde", "T"
                            elif low in ("n","noche","noches"):
                                tlong, code = "Noches", "N"
                            elif low in ("d","descanso","libre"):
                                tlong, code = "Descanso", "D"
                            elif low in ("", "nan"):
                                tlong, code = "", ""
                            else:
                                tlong, code = t, ""

                        rows.append({
                            "Hotel": sh_name,
                            "SemanaNum": semana_num,
                            "SemanaInicio": week_start,
                            "Empleado": emp,
                            "EmpRow": rr,            # respeta orden Excel
                            "Dia": dname,
                            "Fecha": day_dates.get(dname, ""),
                            "Turno": code,
                            "TurnoLargo": tlong,
                            "NameColorC": name_hex,
                            "SubColorL": sub_hex
                        })
                    rr += 1
                r = rr + 1
            else:
                r += 1
    return rows

def build_meta(rows):
    nights = Counter()
    for r in rows:
        if is_night(r.get("Turno"), r.get("TurnoLargo")):
            nights[r["Empleado"]] += 1
    return {"nights_per_employee": dict(nights)}

def main():
    ap = argparse.ArgumentParser(description="Convierte Excel de turnos a JSON (bloques SEMANA, color RGB sólido).")
    ap.add_argument("excel", help="Ruta al Excel (ej: 'TURNOS  v.5.2.xlsx')")
    ap.add_argument("json_out", help="Ruta JSON de salida")
    args = ap.parse_args()

    rows = parse_excel(args.excel)
    meta = build_meta(rows)
    out = {"rows": rows, "meta": meta}
    with open(args.json_out, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"✓ Generado: {args.json_out}  ({len(rows)} filas)")
    print(f"  Noches totales (archivo): {sum(meta['nights_per_employee'].values())}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python excel_a_turnos_json.py input.xlsx output.json")
        sys.exit(1)
    main()
